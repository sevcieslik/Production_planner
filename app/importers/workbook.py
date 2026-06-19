from __future__ import annotations
from dataclasses import dataclass
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook

DISCIPLINE_TOKENS = ('RS', 'GIS', 'PLS')
WORKFLOW_TOKENS = ('capacity', 'allocation', 'allocations', 'demand', 'forecast', 'actual', 'osr', 'leave', 'holiday', 'roster')

@dataclass
class SheetProfile:
    sheet_name: str
    rows: int
    columns: int
    likely_role: str
    detected_disciplines: str
    header_preview: str


def _classify(sheet_name: str, headers: list[str]) -> str:
    text = ' '.join([sheet_name, *headers]).lower()
    if 'osr' in text or 'actual' in text or 'progress' in text:
        return 'OSR actuals/progress source'
    if 'leave' in text or 'holiday' in text or 'absence' in text:
        return 'Availability reducer'
    if 'roster' in text or 'people' in text or 'staff' in text:
        return 'People / roster source'
    if 'alloc' in text:
        return 'Daily/person allocation source'
    if 'capacity' in text or 'demand' in text or any(d.lower() in text for d in DISCIPLINE_TOKENS):
        return 'Weekly discipline demand/capacity source'
    return 'Reference / unknown'


def profile_workbook(uploaded_file) -> list[SheetProfile]:
    data = uploaded_file.getvalue() if hasattr(uploaded_file, 'getvalue') else uploaded_file.read()
    wb = load_workbook(BytesIO(data), read_only=True, data_only=False)
    profiles=[]
    for ws in wb.worksheets:
        preview=[]
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            vals=[str(v).strip() for v in row if v is not None and str(v).strip()]
            if vals:
                preview.extend(vals[:12])
        headers=preview[:20]
        discs=', '.join([d for d in DISCIPLINE_TOKENS if any(d.lower() in h.lower() for h in headers+[ws.title])])
        profiles.append(SheetProfile(ws.title, ws.max_row, ws.max_column, _classify(ws.title, headers), discs, ' | '.join(headers[:12])))
    return profiles


def profiles_to_frame(profiles: list[SheetProfile]) -> pd.DataFrame:
    return pd.DataFrame([p.__dict__ for p in profiles])
